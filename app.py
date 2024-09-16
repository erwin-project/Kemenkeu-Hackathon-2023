from streamlit_multipage import MultiPage
from utils import check_email, check_account, update_json, replace_json
from utils import visualization as vs
from utils import machine_learning as ml
from PIL import Image
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl as pxl
from sklearn.preprocessing import MinMaxScaler
import warnings
warnings.filterwarnings("ignore")


def sign_up(st, **state):
    placeholder = st.empty()

    with placeholder.form("Sign Up"):
        image = Image.open("images/logo_fhas.png")
        st1, st2, st3 = st.columns(3)

        with st2:
            st.image(image)

        st.warning("Please sign up your account!")

        # name_ = state["name"] if "name" in state else ""
        name = st.text_input("Name: ")

        # username_ = state["username"] if "username" in state else ""
        username = st.text_input("Username: ")

        # email_ = state["email"] if "email" in state else ""
        email = st.text_input("Email")

        # password_ = state["password"] if "password" in state else ""
        password = st.text_input("Password", type="password")

        save = st.form_submit_button("Save")

    if save and check_email(email) == "valid email":
        placeholder.empty()
        st.success("Hello " + name + ", your profile has been save successfully")
        MultiPage.save({"name": name,
                        "username": username,
                        "email": email,
                        "password": password,
                        "login": "True",
                        "edit": True})

        update_json(name, username, email, password)

    elif save and check_email(email) == "duplicate email":
        st.success("Hello " + name + ", your profile hasn't been save successfully because your email same with other!")

    elif save and check_email(email) == "invalid email":
        st.success("Hello " + name + ", your profile hasn't been save successfully because your email invalid!")

    else:
        pass


def login(st, **state):
    st.snow()
    # Create an empty container
    placeholder = st.empty()

    try:
        # Insert a form in the container
        with placeholder.form("login"):
            image = Image.open("images/logo_fhas.png")
            st1, st2, st3 = st.columns(3)

            with st2:
                st.image(image)

            st.markdown("#### Login FHAS Website")
            email = st.text_input("Email")
            password = st.text_input("Password", type="password")
            submit = st.form_submit_button("Login")

            st.write("Are you ready registered account in this app? If you don't yet, please sign up your account!")

            name, username, status = check_account(email, password)

        if submit and status == 'register':
            # If the form is submitted and the email and password are correct,
            # clear the form/container and display a success message
            placeholder.empty()
            st.success("Login successful")
            MultiPage.save({"name": name,
                            "username": username,
                            "email": email,
                            "password": password,
                            "login": "True"})

        elif submit and status == 'wrong password':
            st.error("Login failed because your password is wrong!")

        elif submit and status == 'not register':
            st.error("You haven't registered to this app! Please sign up your account!")

        else:
            pass

    except:
        st.error("Please login with your registered email!")


def dashboard(st, **state):
    # Title
    image = Image.open("images/logo_fhas.png")
    st1, st2, st3 = st.columns(3)

    with st2:
        st.image(image)

    st.markdown("<svg width=\"705\" height=\"5\"><line x1=\"0\" y1=\"2.5\" x2=\"705\" y2=\"2.5\" stroke=\"black\" "
                "stroke-width=\"4\" fill=\"black\" /></svg>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"text-align:center;\">Dashboard</h3>", unsafe_allow_html=True)

    path_data = 'data/data_true.xlsx'
    data = pxl.load_workbook(path_data)
    sheet = data.sheetnames

    select = st.selectbox("Please select your data!",
                          sheet)

    dataset = pd.read_excel(path_data,
                            sheet_name=select)

    data_years = dataset.columns.values

    years = st.selectbox("Please select year do you want!",
                         data_years[1:])

    kind = st.selectbox("Please select your chart bar do you want!",
                        ["Horizontal Bar Chart",
                         "Vertical Bar Chart",
                         "Map Geospatial"])

    chart_data = dataset.loc[:, ['Provinsi',
                                 years]]

    chart_datas = pd.melt(chart_data, id_vars=["Provinsi"])

    titles = str("Graph " + " '" + select + "' in " + str(years))

    if kind == "Vertical Bar Chart":
        st.altair_chart(vs.get_bar_vertical(chart_datas,
                                            "Provinsi",
                                            "value",
                                            "variable",
                                            "Provinsi",
                                            "Value",
                                            titles))
    elif kind == "Horizontal Bar Chart":
        st.altair_chart(vs.get_bar_horizontal(chart_datas,
                                              "value",
                                              "Provinsi",
                                              "variable",
                                              "Value",
                                              "Provinsi",
                                              titles))
    elif kind == "Map Geospatial":
        fig, ax = vs.get_chart_map(chart_data,
                                   years,
                                   titles,
                                   'Kementrian Keuangan')

        st.pyplot(fig)


def insight(st, **state):
    # Title
    image = Image.open("images/logo_fhas.png")
    st1, st2, st3 = st.columns(3)

    with st2:
        st.image(image)

    st.markdown("<svg width=\"705\" height=\"5\"><line x1=\"0\" y1=\"2.5\" x2=\"705\" y2=\"2.5\" stroke=\"black\" "
                "stroke-width=\"4\" fill=\"black\" /></svg>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"text-align:center;\">Data Insight</h3>", unsafe_allow_html=True)

    restriction = state["login"]

    if "login" not in state or restriction == "False":
        st.warning("Please login with your registered email!")
        return

    path_data = 'data/data_true.xlsx'
    data = pxl.load_workbook(path_data)
    sheet = data.sheetnames

    dataset1 = pd.read_excel(path_data,
                             sheet_name=sheet[0])

    data_province = dataset1['Provinsi'].values

    label = st.selectbox("Please select province do you want!",
                         data_province)
    i = 0
    for column in sheet:
        datas = pd.read_excel(path_data,
                              sheet_name=column)

        dataset = pd.melt(datas, id_vars=["Provinsi"])
        chart_data = dataset[dataset['Provinsi'] == label]

        titles = str("Graph " + " '" + column + "'")

        if i % 2 == 0:
            st1, st2 = st.columns(2)
            with st1:
                chart_datas = chart_data.loc[:, ["variable",
                                                 "value"]]

                fig, ax = vs.get_bar_vertical_1(chart_datas, titles)

                st.pyplot(fig)

                # fig = vs.get_bar_vertical_2(chart_data, 'variable', 'value', 'Provinsi', 'Years', 'Value', titles)
                #
                # st.altair_chart(fig)

        else:
            with st2:
                chart_datas = chart_data.loc[:, ["variable",
                                                 "value"]]

                fig, ax = vs.get_bar_vertical_1(chart_datas, titles)

                st.pyplot(fig)

                # fig = vs.get_bar_vertical_2(chart_data, 'variable', 'value', 'Provinsi', 'Years', 'Value', titles)
                #
                # st.altair_chart(fig)

        i += 1


def exploratory_data(st, **state):
    # Title
    image = Image.open("images/logo_fhas.png")
    st1, st2, st3 = st.columns(3)

    with st2:
        st.image(image)

    st.markdown("<svg width=\"705\" height=\"5\"><line x1=\"0\" y1=\"2.5\" x2=\"705\" y2=\"2.5\" stroke=\"black\" "
                "stroke-width=\"4\" fill=\"black\" /></svg>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"text-align:center;\">Exploratory Data</h3>", unsafe_allow_html=True)

    restriction = state["login"]

    if "login" not in state or restriction == "False":
        st.warning("Please login with your registered email!")
        return

    path_data = 'data/data_true.xlsx'
    data = pxl.load_workbook(path_data)
    sheet = data.sheetnames

    st1, st2, st3 = st.columns(3)

    with st1:
        select1 = st.selectbox("Please select your data in x-axis!",
                               sheet)

    with st2:
        select2 = st.selectbox("Please select your data in y-axis!",
                               sheet)

    with st3:
        select3 = st.selectbox("Please select your data in xy-axis!",
                               sheet)

    dataset1 = pd.read_excel(path_data,
                             sheet_name=select1)
    dataset2 = pd.read_excel(path_data,
                             sheet_name=select2)
    dataset3 = pd.read_excel(path_data,
                             sheet_name=select3)

    column = list(((set(dataset1.drop('Provinsi', axis=1).columns.values)
                    .intersection(set(dataset2.drop('Provinsi', axis=1).columns.values)))
                   .intersection(set(dataset3.drop('Provinsi', axis=1).columns.values))))

    years = st.selectbox("Please select years do you want!",
                         column)

    data = pd.DataFrame({'Provinsi': dataset1['Provinsi'].values,
                         select1: dataset1[years].values,
                         select2: dataset2[years].values,
                         select3: dataset3[years].values})

    title = str("Exploratory Data in " + str(years) + " " + "\n '" + select1 + "' and '" + select2)

    chart = vs.cross_data(data, select1, select2, select3, title)
    y_pred, score = ml.linear_regression(data[select1], data[select2])

    st.success("The parameter " + " '" + select1 + "' and '" + select2 + "' has correlation " + str(round(score, 2)))
    st.altair_chart(chart)


def efficiency_prediction(st, **state):
    image = Image.open("images/logo_fhas.png")
    st1, st2, st3 = st.columns(3)

    with st2:
        st.image(image)

    st.markdown("<svg width=\"705\" height=\"5\"><line x1=\"0\" y1=\"2.5\" x2=\"705\" y2=\"2.5\" stroke=\"black\" "
                "stroke-width=\"4\" fill=\"black\" /></svg>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"text-align:center;\">Efficiency Prediction</h3>", unsafe_allow_html=True)

    restriction = state["login"]

    if "login" not in state or restriction == "False":
        st.warning("Please login with your registered email!")
        return

    path_data = 'data/data_true.xlsx'
    data = pxl.load_workbook(path_data)
    sheet = data.sheetnames

    data_model = []
    data_pertumbuhan = []
    data_target = []

    for col in sheet:
        if "Efisiensi" in col:
            data_target.append(col)
        else:
            if 'Pertumbuhan' in col:
                data_pertumbuhan.append(col)
            data_model.append(col)

    dataset = pd.read_excel(path_data,
                            sheet_name=sheet[0])
    province = st.selectbox('Please select your province do you want!',
                            dataset['Provinsi'])
    target_efficiency = st.selectbox('Please select your province do you want!',
                                     data_target)

    data_ml, score = ml.convert_data_efficiency(path_data,
                                                target_efficiency)

    dataset_efficiency = pd.read_excel(path_data,
                                       sheet_name=target_efficiency)
    dataset_efficiency[2021], score = ml.linear_regression(dataset_efficiency[2019].values,
                                                           dataset_efficiency[2020].values)

    data_true = pd.melt(dataset_efficiency, id_vars=["Provinsi"])
    chart_data = data_true[data_true['Provinsi'] == province]

    titles = str("Graph " + " '" + target_efficiency + "'")

    chart = vs.get_bar_vertical(chart_data, 'variable', 'value', 'variable', 'Years', 'Value', titles)
    st.altair_chart(chart)


def deployment_model(st, **state):
    image = Image.open("images/logo_fhas.png")
    st1, st2, st3 = st.columns(3)

    with st2:
        st.image(image)

    st.markdown("<svg width=\"705\" height=\"5\"><line x1=\"0\" y1=\"2.5\" x2=\"705\" y2=\"2.5\" stroke=\"black\" "
                "stroke-width=\"4\" fill=\"black\" /></svg>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"text-align:center;\">Deployment Model</h3>", unsafe_allow_html=True)

    restriction = state["login"]

    if "login" not in state or restriction == "False":
        st.warning("Please login with your registered email!")
        return

    path_data = 'data/data_true.xlsx'
    data = pxl.load_workbook(path_data)
    sheet = data.sheetnames

    try:
        st1, st2 = st.columns(2)

        with st1:
            select = st.multiselect('Please select data do you want to build model!',
                                    sheet)

            target = st.selectbox('Please select your target model!',
                                  select)

            i = 0
            for col in select:
                data_true = pd.read_excel(path_data,
                                          sheet_name=col)

                if i == 0:
                    cols = set(data_true.drop('Provinsi', axis=1).columns.values)
                else:
                    cols = cols.intersection(data_true.drop('Provinsi', axis=1).columns.values)

                i += 1

        with st2:
            years_data = list(cols)

            years = st.selectbox('Please select years do you want to build model!',
                                 years_data)

            proj_years = st.selectbox('Please select projection years do you want to predict!',
                                      years_data)

            data_ml = pd.DataFrame({'Provinsi': data_true['Provinsi']})
            data_ml_proj = pd.DataFrame({'Provinsi': data_true['Provinsi']})

            for datas in select:
                data_excel = pd.read_excel(path_data,
                                           sheet_name=datas)

                data_ml[datas] = data_excel[years].values
                data_ml_proj[datas] = data_excel[proj_years].values

        st.success('Your data has been successfully saved')
        st.dataframe(data_ml)

        st.markdown("<h3 style=\"text-align:center;\">Data Cleaning and Transformation<h3>", unsafe_allow_html=True)

        st3, st4 = st.columns(2)

        with st3:
            check = st.radio('Do you want to cleaning data?',
                             ['No', 'Yes'])

            if check == 'Yes':
                data_ml.dropna(inplace=True)
                # st.write("Table NaN Values")
                # st.write(np.transpose(data_ml.isnull().sum()))

        with st4:
            transform = st.radio('Do you want to transform data with MaxMinScaler?',
                                 ['No', 'Yes'])

            if transform == 'Yes':
                scaler = MinMaxScaler(feature_range=(0, 1))

                for data_col in data_ml.columns:
                    if data_col != 'Provinsi':
                        data_ml[data_col] = scaler.fit_transform(data_ml[data_col].values.reshape(-1, 1))
                        data_ml_proj[data_col] = scaler.fit_transform(data_ml_proj[data_col].values.reshape(-1, 1))

        if check == 'Yes' and transform == 'Yes':
            st.success('Your data has been successfully cleaned and transformed')

        st.dataframe(data_ml)

        st.markdown("<h3 style=\"text-align:center;\">Building Model<h3>", unsafe_allow_html=True)

        models = st.selectbox('Please select your kind model machine learning!',
                              ['Supervised Learning',
                               'Unsupervised Learning'])

        box = []

        if models == 'Supervised Learning':
            box = ['Linear Regression',
                   'Bayesian Ridge Regression',
                   'SVR',
                   'Decision Tree Regression']

        elif models == 'Unsupervised Learning':
            box = ['LSTM',
                   'ANN']

        kind_model = st.selectbox('Please select your model machine learning!',
                                  box)

        if models == 'Supervised Learning':
            chart1, chart2, score, dataset = ml.supervised_learning(kind_model,
                                                                    scaler,
                                                                    data_ml,
                                                                    data_ml_proj,
                                                                    target,
                                                                    years,
                                                                    proj_years)
        elif models == "Unsupervised Learning":
            chart1, chart2, score, dataset = ml.unsupervised_learning(kind_model,
                                                                      scaler,
                                                                      data_ml,
                                                                      data_ml_proj,
                                                                      target,
                                                                      years,
                                                                      proj_years)

        st.success('Your model has accuracy ' + str(round(score, 2)))

        st.altair_chart(chart1)
        st.altair_chart(chart2)

    except:
        st.error('First, please select data and years do you want!')


def report(st, **state):
    # Title
    image = Image.open("images/logo_fhas.png")
    st1, st2, st3 = st.columns(3)

    with st2:
        st.image(image)

    st.markdown("<svg width=\"705\" height=\"5\"><line x1=\"0\" y1=\"2.5\" x2=\"705\" y2=\"2.5\" stroke=\"black\" "
                "stroke-width=\"4\" fill=\"black\" /></svg>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"text-align:center;\">Messages Report</h3>", unsafe_allow_html=True)

    restriction = state["login"]

    if "login" not in state or restriction == "False":
        st.warning("Please login with your registered email!")
        return

    placeholder = st.empty()

    with placeholder.form("Message"):
        email = st.text_input("Email")
        text = st.text_area("Messages")
        submit = st.form_submit_button("Send")

    if submit and check_email(email) == "valid email" or check_email(email) == "duplicate email":
        placeholder.empty()
        st.success("Before your message will be send, please confirm your messages again!")
        vals = st.write("<form action= 'https://formspree.io/f/xeqdqdon' "
                        "method='POST'>"
                        "<label> Email: <br> <input type='email' name='email' value='" + str(email) +
                        "'style='width:705px; height:50px;'></label>"
                        "<br> <br>"
                        "<label> Message: <br> <textarea name='Messages' value='" + str(text) +
                        "'style='width:705px; height:200px;'></textarea></label>"
                        "<br> <br>"
                        "<button type='submit'>Confirm</button>"
                        "</form>", unsafe_allow_html=True)

        if vals is not None:
            st.success("Your messages has been send successfully!")

    elif submit and check_email(email) == "invalid email":
        st.success("Your message hasn't been send successfully because email receiver not in list")

    else:
        pass


def account(st, **state):
    # Title
    image = Image.open("images/logo_fhas.png")
    st1, st2, st3 = st.columns(3)

    with st2:
        st.image(image)

    st.markdown("<svg width=\"705\" height=\"5\"><line x1=\"0\" y1=\"2.5\" x2=\"705\" y2=\"2.5\" stroke=\"black\" "
                "stroke-width=\"4\" fill=\"black\" /></svg>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"text-align:center;\">Account Setting</h3>", unsafe_allow_html=True)

    restriction = state["login"]
    password = state["password"]

    if "login" not in state or restriction == "False":
        st.warning("Please login with your registered email!")
        return

    placeholder = st.empty()

    st.write("Do you want to edit your account?")
    edited = st.button("Edit")
    state["edit"] = np.invert(edited)

    old_email = state['email']

    with placeholder.form("Account"):
        name_ = state["name"] if "name" in state else ""
        name = st.text_input("Name", placeholder=name_, disabled=state["edit"])

        username_ = state["username"] if "username" in state else ""
        username = st.text_input("Username", placeholder=username_, disabled=state["edit"])

        email_ = state["email"] if "email" in state else ""
        email = st.text_input("Email", placeholder=email_, disabled=state["edit"])

        if edited:
            current_password = st.text_input("Old Password", type="password", disabled=state["edit"])
        else:
            current_password = password

        # current_password_ = state["password"] if "password" in state else ""
        new_password = st.text_input("New Password", type="password", disabled=state["edit"])

        save = st.form_submit_button("Save")

    if save and current_password == password:
        st.success("Hi " + name + ", your profile has been update successfully")
        MultiPage.save({"name": name,
                        "username": username,
                        "email": email,
                        "password": new_password,
                        "edit": True})

        replace_json(name, username, old_email, email, new_password)

    elif save and current_password != password:
        st.success("Hi " + name + ", your profile hasn't been update successfully because your current password"
                                  " doesn't match!")

    elif save and check_email(email) == "invalid email":
        st.success("Hi " + name + ", your profile hasn't been update successfully because your email invalid!")

    else:
        pass


def logout(st, **state):
    st.success("Your account has been log out from this app")
    MultiPage.save({"login": "False"})


app = MultiPage()
app.st = st

app.navbar_name = "Menu"
app.navbar_style = "VerticalButton"

app.hide_menu = False
app.hide_navigation = True

app.add_app("Sign Up", sign_up)
app.add_app("Login", login)
app.add_app("Dashboard", dashboard)
app.add_app("Data Insight", insight)
app.add_app("Exploratory Data", exploratory_data)
app.add_app("Efficiency Prediction", efficiency_prediction)
app.add_app("Deployment Model", deployment_model)
app.add_app("Report", report)
app.add_app("Account Setting", account)
app.add_app("Logout", logout)

app.run()
